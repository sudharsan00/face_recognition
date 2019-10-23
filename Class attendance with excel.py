#!/usr/bin/env python
# coding: utf-8

#                                              Attendance with Face Recognition

# In[1]:


# import

import numpy as np
import datetime as dt
import cv2
import os,sys
import face_recognition as fr
import openpyxl
from openpyxl import Workbook, load_workbook
from PIL import Image
import pyttsx3


# In[2]:


# SpreadSheet Creation

date_var = dt.datetime.now()
date = date_var.strftime("%x")
excel_path = "D:/Data science/Attendance/Excel/Attendance.xlsx"
wb = openpyxl.load_workbook(excel_path) 
excel_sheet_daily=wb.get_sheet_by_name('Daily_attendance')
xl_sheet = wb.get_sheet_by_name('period_attendance')
for iterate4 in excel_sheet_daily.iter_cols(max_row =1, min_col=3, max_col=8):
    for cells in iterate4:
        split_cell_partial = str(cells).split(".")[1]
        split_cell = split_cell_partial.split(">")[0]
        print(split_cell)
        cell = excel_sheet_daily[split_cell]
        check_value_exist = cell.value
        print(check_value_exist)
    if check_value_exist == date:
        break
    elif check_value_exist == None:
        cell = excel_sheet_daily[split_cell]
        cell.value = date
        wb.save(excel_path)
        print(cell.value)
        for itearte5 in xl_sheet['C2:G58']:
            for iterate6 in itearte5:
                iterate6.value = None
                wb.save(excel_path)
        break


# In[3]:


# Load Train Dataset

get_ipython().run_line_magic('cd', '"D:\\Data science\\Attendance\\Image dataset\\train"')
loaded_images_original = []
original_path = "D:/Data science/Attendance/Image dataset/train"
make_list_original = os.listdir(original_path)
for filename_1 in make_list_original:
    loaded_images_original.append(filename_1)
print(loaded_images_original)

train_path = "D:/Data science/Attendance/Image dataset/Trainfaces"
for iterate1 in loaded_images_original:
    print(iterate1)
    image = cv2.imread(iterate1)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    face_cascade = cv2.CascadeClassifier('C:\\Users\\user\\Anaconda3\\pkgs\\opencv-4.1.1-py37h6afde12_1\\Library\\etc\\haarcascades\\haarcascade_frontalface_default.xml')
    faces = face_cascade.detectMultiScale(gray, scaleFactor= 1.3,minNeighbors = 5)
    for (x, y, width, height) in faces:
        w = x+width
        h = y+height
        cv2.rectangle(gray, (x, y), (w,h), (255, 255, 255), 1)
        roi_gray = gray[y:h, x:w]
        cv2.imwrite(os.path.join(train_path, iterate1),roi_gray)


# In[4]:


# Load Test Dataset

get_ipython().run_line_magic('cd', '"D:\\Data science\\Attendance\\Image dataset\\test"')
loaded_images_test = []
test_path = "D:/Data science/Attendance/Image dataset/test"
make_list_test = os.listdir(test_path)
for filename_2 in make_list_test:
    loaded_images_test.append(filename_2)
print(loaded_images_test)
test_face_path = "D:/Data science/Attendance/Image dataset/testfaces"
lent1 = 0
c = []
d = []
appen = []
for iterate11 in loaded_images_test:
    print(iterate11)
    test_image = cv2.imread(iterate11)
    test_gray = cv2.cvtColor(test_image, cv2.COLOR_BGR2GRAY)
    test_face_cascade = cv2.CascadeClassifier('C:\\Users\\user\\Anaconda3\\pkgs\\opencv-4.1.1-py37h6afde12_1\\Library\\etc\\haarcascades\\haarcascade_frontalface_default.xml')
    test_faces = test_face_cascade.detectMultiScale(test_gray, scaleFactor= 1.3,minNeighbors = 5)
    lent = len(test_faces)
    print(f"length of testface : {lent}")
    if lent1 == 0:
        lent1 = lent1 +len(test_faces)
        print(lent1)
        a = list(range(0,lent))
        print(a)
    else:
        lent1 = lent1 +len(test_faces)
        print(lent1)
        var = a[-1]+1
        print(f'var : {var}')
        a = list(range(var,lent1))
        print(a)
    for b in a:
        jpg= ".jpg"
        new_face = str(b) + jpg
        c.append(new_face)
    for (tx, ty, twidth, theight) in test_faces:
        tw = tx+twidth
        th = ty+theight
        e = tx,ty,tw,th
        d.append(e)
    for g in a:
        for f in d[g]:
            img = Image.open(iterate11).convert('L')
            img2 = img.crop(d[g])
            img2.save(os.path.join(test_face_path,str(c[g])))
            break


# In[5]:


# Select Period 

get_ipython().run_line_magic('cd', '"D:\\Data science\\Attendance\\Excel"')
period = input("Enter period :")
print(xl_sheet.title)
train_image_paths=[os.path.join(train_path,fe) for fe in os.listdir(train_path)] 
test_image_paths = [os.path.join(test_face_path,de) for de in os.listdir(test_face_path)]
test_oimage_paths = [os.path.join(test_path,de) for de in os.listdir(test_path)]
Id1 = []
Id2 = []
def select_per(period):
    if int(period)==1:
        select_period= 3
        print_in_sheet(select_period)
    elif int(period)==2:
        select_period= 4
        print_in_sheet(select_period)
    elif int(period)==3:
        select_period=5
        print_in_sheet(select_period)
    elif int(period)==4:
        select_period=6
        print_in_sheet(select_period)
    elif int(period)==5:
        select_period=7
        print_in_sheet(select_period)
    else:
        print("Period out of bound")

# Compare Faces With Training Dataset And Enter In Excel Sheet

def print_in_sheet(select_period):
    for iterate3 in test_image_paths:
        Id2 = int(os.path.split(iterate3)[-1].split(".")[0])
        print(Id2)
        for iterate2 in train_image_paths:
            Id1 = int(os.path.split(iterate2)[-1].split(".")[0])
            known_image = fr.load_image_file(iterate2)
            unknown_image = fr.load_image_file(iterate3)
            known_encoding = fr.face_encodings(known_image)[0]
            unknown_encoding = fr.face_encodings(unknown_image)[0]
            results = fr.compare_faces([known_encoding], unknown_encoding)
            if results[0] == True:
                if Id1 == 1:
                    print("Sathyaseelan")
                    c1= xl_sheet.cell(row=2,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 2:
                    print("Sekeshwaran")
                    c1= xl_sheet.cell(row=3,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 3:
                    print("Sharkash")
                    c1= xl_sheet.cell(row=4,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 4 :
                    print("Sharmi")
                    c1= xl_sheet.cell(row=5,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 5:
                    print("Sharon")
                    c1= xl_sheet.cell(row=6,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 6:
                    print("Sashmitha")
                    c1= xl_sheet.cell(row=7,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 7:
                    print("Sheik")
                    c1= xl_sheet.cell(row=8,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 8:
                    print("Shijo")
                    c1= xl_sheet.cell(row=9,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 9:
                    print("Sibi")
                    c1= xl_sheet.cell(row=10,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 10:
                    print("Siva")
                    c1= xl_sheet.cell(row=11,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 11:
                    print("Sivanesan")
                    c1= xl_sheet.cell(row=12,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 12:
                    print("Sivaraj")
                    c1= xl_sheet.cell(row=13,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 13:
                    print("Sivaraman G")
                    c1= xl_sheet.cell(row=14,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 14:
                    print("Sivaraman")
                    c1= xl_sheet.cell(row=15,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 15:
                    print("Sivashankari")
                    c1= xl_sheet.cell(row=16,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 16:
                    print("Siva subramaniyan")
                    c1= xl_sheet.cell(row=17,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 17:
                    print("Sowndarya")
                    c1= xl_sheet.cell(row=18,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 18:
                    print("Sree vignesh")
                    c1= xl_sheet.cell(row=19,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 19:
                    print("Skandha moorthy")
                    c1= xl_sheet.cell(row=20,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                '''elif Id1 == 20:
                    print("Srinidhi")
                    c1= xl_sheet.cell(row=21,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 21:
                    print("Sriram")
                    c1= xl_sheet.cell(row=22,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 22:
                    print("Subaranjani")
                    c1= xl_sheet.cell(row=23,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 23:
                    print("Subasri")
                    c1= xl_sheet.cell(row=24,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 23:
                    print("Subraja")
                    c1= xl_sheet.cell(row=25,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 24:
                    print("Sudharsan")
                    c1= xl_sheet.cell(row=26,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 25:
                    print("Suganth")
                    c1= xl_sheet.cell(row=27,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 26:
                    print("SUjan patrick ezhbrich")
                    c1= xl_sheet.cell(row=28,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 27:
                    print("Surendar")
                    c1= xl_sheet.cell(row=29,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 28:
                    print("Suresh")
                    c1= xl_sheet.cell(row=30,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 29:
                    print("Sushma")
                    c1= xl_sheet.cell(row=31,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 30:
                    print("Tamil azhagan")
                    c1= xl_sheet.cell(row=32,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 31:
                    print("Thanushwar")
                    c1= xl_sheet.cell(row=33,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 32:
                    print("Thilothini")
                    c1= xl_sheet.cell(row=34,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 33:
                    print("Uthrapathi")
                    c1= xl_sheet.cell(row=35,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 34:
                    print("Vasanth")
                    c1= xl_sheet.cell(row=36,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 35:
                    print("Veeramani")
                    c1= xl_sheet.cell(row=37,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 36:
                    print("Vel kumar")
                    c1= xl_sheet.cell(row=38,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 37:
                    print("Vellaisamy")
                    c1= xl_sheet.cell(row=39,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 38:
                    print("Vetrivel")
                    c1= xl_sheet.cell(row=40,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 39:
                    print("Vignesh R")
                    c1= xl_sheet.cell(row=41,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 40:
                    print("Vignesh S 45")
                    c1= xl_sheet.cell(row=42,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 41:
                    print("Vignesh S 46")
                    c1= xl_sheet.cell(row=43,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 42:
                    print("Vignesh S 47")
                    c1= xl_sheet.cell(row=44,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 43:
                    print("Vignesh P")
                    c1= xl_sheet.cell(row=45,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 45:
                    print("Vigneshwaran")
                    c1= xl_sheet.cell(row=46,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 46:
                    print("Vigneshwarai")
                    c1= xl_sheet.cell(row=47,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 47:
                    print("Vijay babu")
                    c1= xl_sheet.cell(row=48,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 48:
                    print("Vijaykumar")
                    c1= xl_sheet.cell(row=49,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 49:
                    print("Vijay anand")
                    c1= xl_sheet.cell(row=50,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 50:
                    print("Vinoth Kumar")
                    c1= xl_sheet.cell(row=51,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 51:
                    print("Vishnu prasath")
                    c1= xl_sheet.cell(row=52,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 52:
                    print("Yamini")
                    c1= xl_sheet.cell(row=53,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 53:
                    print("Yokeshwaran 57")
                    c1= xl_sheet.cell(row=54,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 54:
                    print("Yokeswaran B")
                    c1= xl_sheet.cell(row=55,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 55:
                    print("Kavi priya")
                    c1= xl_sheet.cell(row=56,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 56:
                    print("Rakesh")
                    c1= xl_sheet.cell(row=57,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break
                elif Id1 == 57:
                    print("Abirooph")
                    c1= xl_sheet.cell(row=58,column=int(select_period))
                    c1.value=1
                    wb.save(excel_path)
                    break'''
select_per(period)


# In[6]:


#Get The Name Of The Absentee

abb_arr = []
whatsapp = []
if int(period) ==1:
    selec_cells= 'C2:C12'
elif int(period) ==2:
    selec_cells= 'D2:D12'
elif int(period) == 3:
    selec_cells= 'E2:E12'
elif int(period) == 4:
    selec_cells= 'F2:F12'
elif int(period) ==5:
    selec_cells= 'G2:G12'
for iterate7 in xl_sheet[selec_cells]:
    for iterate8 in iterate7:
        if iterate8.value == None:   
            split_speech_par = str(iterate8).split(".")[1]
            split_speech = split_speech_par.split(">")[0]
            if int(period) ==1:
                split_add = split_speech.replace("C","B")
                split_add_b = xl_sheet[split_add].value
                abb_arr.append(split_add_b)
                print(split_speech)
                whatsapp_replace = split_add.replace("B","M")
                whatsapp_replace_b = xl_sheet[whatsapp_replace].value
                whatsapp.append(whatsapp_replace_b)
            elif int(period) ==2:
                split_add = split_speech.replace("D","B")
                split_add_b = xl_sheet[split_add].value
                abb_arr.append(split_add_b)
                whatsapp_replace = split_add.replace("B","M")
                whatsapp_replace_b = xl_sheet[whatsapp_replace].value
                whatsapp.append(whatsapp_replace_b)
            elif int(period) ==3:
                split_add = split_speech.replace("E","B")
                split_add_b = xl_sheet[split_add].value
                abb_arr.append(split_add_b)
                whatsapp_replace = split_add.replace("B","M")
                whatsapp_replace_b = xl_sheet[whatsapp_replace].value
                whatsapp.append(whatsapp_replace_b)
            elif int(period) == 4:
                split_add = split_speech.replace("F","B")
                split_add_b = xl_sheet[split_add].value
                abb_arr.append(split_add_b)
                whatsapp_replace = split_add.replace("B","M")
                whatsapp_replace_b = xl_sheet[whatsapp_replace].value
                whatsapp.append(whatsapp_replace_b)
            elif int(period) == 5:
                split_add = split_speech.replace("G","B")
                split_add_b = xl_sheet[split_add].value
                abb_arr.append(split_add_b)
                whatsapp_replace = split_add.replace("B","M")
                whatsapp_replace_b = xl_sheet[whatsapp_replace].value
                whatsapp.append(whatsapp_replace_b)
print(whatsapp)


# In[7]:


# Spell The Name of The Absentee

speech = pyttsx3.init()
voices = speech.getProperty('voices')
speech.setProperty('voices', voices[1].id)
for iterate9 in abb_arr:
    print(iterate9)
    speech.say(iterate9)
    speech.runAndWait()
speech.say(f"are absent on {date} period {period}, Please check with tutor if not.")
speech.runAndWait()
speech.stop()


# In[10]:


# Send Notification Through Whatsapp

print("Should I notify the Absentee ? \nHit y if YES \nHit n If NO")
def whatsapp():
    whether_notify = input()
    if whether_notify == 'y' or whether_notify == 'Y':
        from selenium import webdriver
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support.ui import Select
        from selenium.webdriver.common.by import By
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.chrome.options import Options
        import time
    
        driver = webdriver.Chrome('C:/Users/user/Downloads/chromedriver_win32/chromedriver.exe') 
        driver.get("https://web.whatsapp.com/") 

        input('Hit Enter only after scanning the qr code')
        print('logged in')
        msg = f'You are marked absent on {date} for period {period}'
        for contact  in whatsapp:    
            inp_xpath_search = "//input[@title='Search or start new chat']"
            input_box_search = WebDriverWait(driver,50).until(lambda driver: driver.find_element_by_xpath(inp_xpath_search))
            input_box_search.click()
            time.sleep(2)
            input_box_search.send_keys(contact)
            time.sleep(2)

            selected_contact = driver.find_element_by_xpath("//span[@title='"+contact+"']")
            selected_contact.click()

            inp_xpath = '//div[@class="_3u328 copyable-text selectable-text"][@contenteditable="true"][@data-tab="1"]'
            input_box = driver.find_element_by_xpath(inp_xpath)
            time.sleep(2)
            input_box.send_keys(msg + Keys.ENTER)
            time.sleep(2)
    elif whether_notify == 'n' or whether_notify == 'N':
        print('Notification not sent')
    else:
        print('Invalid input')
        whatsapp()
whatsapp()


# In[9]:


# Clear The Files 

for rem in train_image_paths:
    os.remove(os.path.join(train_path,rem))
for rem1 in test_image_paths:
    os.remove(os.path.join(test_face_path,rem1))
#for rem2 in test_oimage_paths:
    #os.remove(os.path.join(test_path,rem2))

